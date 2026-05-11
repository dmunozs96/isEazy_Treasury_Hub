import io
import uuid
from datetime import date, datetime
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.bank_account import BankAccount
from app.models.classification import (
    CategoryTaxonomy,
    ClassificationSource,
    MovementClassification,
)
from app.models.company import Company
from app.models.movement import Movement
from app.schemas.classification import SingleClassifyResponse
from app.schemas.common import PaginatedResponse
from app.schemas.movement import ClassificationInfo, ClassificationOverride, MovementResponse
from app.services.classification.engine import classify_movement, load_active_rules

router = APIRouter(prefix="/movements", tags=["movements"])

_SORT_COLS = {
    "value_date": Movement.value_date,
    "amount": Movement.amount,
    "accounting_date": Movement.accounting_date,
}


def _base_query():
    return (
        select(
            Movement.id,
            Movement.company_id,
            Company.short_name.label("company_short_name"),
            Movement.bank_account_id,
            BankAccount.bank_name.label("bank_name"),
            Movement.import_batch_id,
            Movement.value_date,
            Movement.accounting_date,
            Movement.amount,
            Movement.currency,
            Movement.balance_after,
            Movement.description,
            Movement.counterpart_name,
            Movement.counterpart_iban,
            Movement.reference,
            Movement.is_intercompany,
            Movement.created_at,
            Movement.is_deleted,
            MovementClassification.category_code.label("mc_category_code"),
            MovementClassification.subcategory_code.label("mc_subcategory_code"),
            MovementClassification.source.label("mc_source"),
            MovementClassification.is_confirmed.label("mc_is_confirmed"),
            MovementClassification.classified_at.label("mc_classified_at"),
            MovementClassification.override_reason.label("mc_override_reason"),
            CategoryTaxonomy.name.label("cat_name"),
            CategoryTaxonomy.cash_flow_section.label("cat_section"),
        )
        .select_from(Movement)
        .outerjoin(Company, Movement.company_id == Company.id)
        .outerjoin(BankAccount, Movement.bank_account_id == BankAccount.id)
        .outerjoin(MovementClassification, Movement.id == MovementClassification.movement_id)
        .outerjoin(CategoryTaxonomy, MovementClassification.category_code == CategoryTaxonomy.code)
        .where(Movement.is_deleted.is_(False))
    )


def _apply_filters(stmt, company_id, bank_account_id, date_from, date_to, category_code, amount_min, amount_max, search):
    if company_id:
        stmt = stmt.where(Movement.company_id == company_id)
    if bank_account_id:
        stmt = stmt.where(Movement.bank_account_id == bank_account_id)
    if date_from:
        stmt = stmt.where(Movement.value_date >= date_from)
    if date_to:
        stmt = stmt.where(Movement.value_date <= date_to)
    if category_code:
        stmt = stmt.where(MovementClassification.category_code == category_code)
    if amount_min is not None:
        stmt = stmt.where(Movement.amount >= amount_min)
    if amount_max is not None:
        stmt = stmt.where(Movement.amount <= amount_max)
    if search:
        p = f"%{search}%"
        stmt = stmt.where(
            or_(
                Movement.description.ilike(p),
                Movement.counterpart_name.ilike(p),
                Movement.reference.ilike(p),
            )
        )
    return stmt


def _row_to_response(row) -> MovementResponse:
    classification = None
    if row.mc_category_code:
        classification = ClassificationInfo(
            category_code=row.mc_category_code,
            category_name=row.cat_name,
            cash_flow_section=row.cat_section.value if row.cat_section else None,
            subcategory_code=row.mc_subcategory_code,
            source=row.mc_source,
            is_confirmed=row.mc_is_confirmed,
            classified_at=row.mc_classified_at,
            override_reason=row.mc_override_reason,
        )
    return MovementResponse(
        id=row.id,
        company_id=row.company_id,
        company_short_name=row.company_short_name,
        bank_account_id=row.bank_account_id,
        bank_name=row.bank_name,
        import_batch_id=row.import_batch_id,
        value_date=row.value_date,
        accounting_date=row.accounting_date,
        amount=row.amount,
        currency=row.currency,
        balance_after=row.balance_after,
        description=row.description,
        counterpart_name=row.counterpart_name,
        counterpart_iban=row.counterpart_iban,
        reference=row.reference,
        is_intercompany=row.is_intercompany,
        created_at=row.created_at,
        is_deleted=row.is_deleted,
        classification=classification,
    )


@router.get("/", response_model=PaginatedResponse[MovementResponse])
async def list_movements(
    company_id: uuid.UUID | None = Query(None),
    bank_account_id: uuid.UUID | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    category_code: str | None = Query(None),
    amount_min: Decimal | None = Query(None),
    amount_max: Decimal | None = Query(None),
    search: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    sort: str = Query("value_date"),
    order: str = Query("desc"),
    db: AsyncSession = Depends(get_db),
):
    # Count (join only what's needed for filter correctness)
    count_stmt = (
        select(func.count(Movement.id))
        .select_from(Movement)
        .outerjoin(MovementClassification, Movement.id == MovementClassification.movement_id)
        .where(Movement.is_deleted.is_(False))
    )
    count_stmt = _apply_filters(
        count_stmt, company_id, bank_account_id, date_from, date_to,
        category_code, amount_min, amount_max, search,
    )
    total: int = (await db.execute(count_stmt)).scalar_one()

    # Data
    stmt = _apply_filters(
        _base_query(), company_id, bank_account_id, date_from, date_to,
        category_code, amount_min, amount_max, search,
    )
    sort_col = _SORT_COLS.get(sort, Movement.value_date)
    stmt = stmt.order_by(sort_col.desc() if order == "desc" else sort_col.asc())
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)

    rows = (await db.execute(stmt)).all()
    pages = max(1, (total + page_size - 1) // page_size)

    return PaginatedResponse(
        items=[_row_to_response(r) for r in rows],
        total=total,
        page=page,
        page_size=page_size,
        pages=pages,
    )


# /export must be declared before /{movement_id} to avoid route shadowing
@router.get("/export")
async def export_movements(
    company_id: uuid.UUID | None = Query(None),
    bank_account_id: uuid.UUID | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    category_code: str | None = Query(None),
    amount_min: Decimal | None = Query(None),
    amount_max: Decimal | None = Query(None),
    search: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
):
    stmt = _apply_filters(
        _base_query(), company_id, bank_account_id, date_from, date_to,
        category_code, amount_min, amount_max, search,
    )
    stmt = stmt.order_by(Movement.value_date.desc()).limit(10_000)
    rows = (await db.execute(stmt)).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Treasury Ledger"

    headers = [
        "Fecha Valor", "Empresa", "Banco", "Descripción",
        "Importe (EUR)", "Categoría", "Saldo Posterior", "Referencia", "Contraparte",
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            row.value_date.isoformat() if row.value_date else "",
            row.company_short_name or "",
            row.bank_name or "",
            row.description or "",
            float(row.amount) if row.amount is not None else 0.0,
            row.mc_category_code or "UNCLASSIFIED",
            float(row.balance_after) if row.balance_after is not None else None,
            row.reference or "",
            row.counterpart_name or "",
        ])

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().isoformat()
    filename = f"treasury_ledger_{today}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{movement_id}", response_model=MovementResponse)
async def get_movement(movement_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    stmt = _base_query().where(Movement.id == movement_id)
    row = (await db.execute(stmt)).one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Movement not found")
    return _row_to_response(row)


@router.patch("/{movement_id}/category", response_model=MovementResponse)
async def override_category(
    movement_id: uuid.UUID,
    body: ClassificationOverride,
    db: AsyncSession = Depends(get_db),
):
    # Verify movement exists
    mv = (
        await db.execute(
            select(Movement).where(
                Movement.id == movement_id, Movement.is_deleted.is_(False)
            )
        )
    ).scalar_one_or_none()
    if not mv:
        raise HTTPException(status_code=404, detail="Movement not found")

    # Validate category code
    cat = (
        await db.execute(
            select(CategoryTaxonomy).where(
                CategoryTaxonomy.code == body.category_code,
                CategoryTaxonomy.is_active.is_(True),
            )
        )
    ).scalar_one_or_none()
    if not cat:
        raise HTTPException(status_code=422, detail=f"Unknown category code: {body.category_code}")

    # Create or update classification
    existing = (
        await db.execute(
            select(MovementClassification).where(
                MovementClassification.movement_id == movement_id
            )
        )
    ).scalar_one_or_none()

    if existing:
        existing.previous_category_code = existing.category_code
        existing.category_code = body.category_code
        existing.subcategory_code = body.subcategory_code
        existing.source = ClassificationSource.MANUAL
        existing.classified_by = "user"
        existing.classified_at = datetime.utcnow()
        existing.override_reason = body.override_reason
        existing.is_confirmed = True
    else:
        db.add(
            MovementClassification(
                movement_id=movement_id,
                category_code=body.category_code,
                subcategory_code=body.subcategory_code,
                source=ClassificationSource.MANUAL,
                classified_by="user",
                classified_at=datetime.utcnow(),
                override_reason=body.override_reason,
                is_confirmed=True,
            )
        )

    await db.commit()

    row = (await db.execute(_base_query().where(Movement.id == movement_id))).one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Movement not found after update")
    return _row_to_response(row)


@router.post("/{movement_id}/classify", response_model=SingleClassifyResponse)
async def classify_single(movement_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    mv = (
        await db.execute(
            select(Movement).where(Movement.id == movement_id, Movement.is_deleted.is_(False))
        )
    ).scalar_one_or_none()
    if not mv:
        raise HTTPException(status_code=404, detail="Movement not found")

    # Preserve manual overrides
    existing = (
        await db.execute(
            select(MovementClassification).where(MovementClassification.movement_id == movement_id)
        )
    ).scalar_one_or_none()
    if existing and existing.source == ClassificationSource.MANUAL:
        return SingleClassifyResponse(
            movement_id=movement_id,
            category_code=existing.category_code,
            rule_id=existing.rule_id,
            source=existing.source.value,
        )

    rules = await load_active_rules(db)
    result = classify_movement(mv, rules)

    if existing:
        existing.category_code = result.category_code
        existing.source = result.source
        existing.rule_id = result.rule_id
        existing.confidence = result.confidence
        existing.classified_at = datetime.utcnow()
        existing.classified_by = "system"
        existing.is_confirmed = True
    else:
        db.add(
            MovementClassification(
                movement_id=movement_id,
                category_code=result.category_code,
                source=result.source,
                rule_id=result.rule_id,
                confidence=result.confidence,
                classified_by="system",
                is_confirmed=True,
            )
        )

    await db.commit()

    return SingleClassifyResponse(
        movement_id=movement_id,
        category_code=result.category_code,
        rule_id=result.rule_id,
        source=result.source.value,
    )
