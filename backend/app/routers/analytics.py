from __future__ import annotations

import io
import uuid
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.bank_account import BankAccount
from app.models.classification import CategoryTaxonomy, MovementClassification
from app.models.company import Company
from app.models.import_batch import ImportBatch, ImportStatus
from app.models.intercompany import IntercompanyMatch, MatchStatus
from app.models.movement import Movement
from app.schemas.analytics import (
    AccountImportStatus,
    BalanceReconciliation,
    CashFlowPeriod,
    CashFlowRow,
    CashFlowSectionSummary,
    CashFlowStatement,
    CompanyCashPosition,
    ConsistencyReport,
    DataQualityWarning,
    DashboardSummary,
    ImportCoverageStatus,
    UnclassifiedRateWarning,
    WeeklyCashFlow,
)

router = APIRouter(prefix="/analytics", tags=["analytics"])


def _iso_week_label(d: date) -> str:
    iso = d.isocalendar()
    return f"W{iso.week:02d} {iso.year}"


def _last_monday(today: date) -> date:
    return today - timedelta(days=today.weekday())


def _month_start(d: date) -> date:
    return date(d.year, d.month, 1)


def _month_end(d: date) -> date:
    return date(d.year + 1, 1, 1) - timedelta(days=1) if d.month == 12 else date(d.year, d.month + 1, 1) - timedelta(days=1)


def _periods_between(date_from: date, date_to: date, granularity: str) -> list[CashFlowPeriod]:
    periods: list[CashFlowPeriod] = []
    if granularity == "monthly":
        current = _month_start(date_from)
        while current <= date_to:
            period_end = min(_month_end(current), date_to)
            periods.append(
                CashFlowPeriod(
                    key=current.isoformat(),
                    label=current.strftime("%b %Y"),
                    start_date=current,
                    end_date=period_end,
                )
            )
            current = date(current.year + 1, 1, 1) if current.month == 12 else date(current.year, current.month + 1, 1)
        return periods

    current = _last_monday(date_from)
    while current <= date_to:
        period_end = min(current + timedelta(days=6), date_to)
        periods.append(
            CashFlowPeriod(
                key=current.isoformat(),
                label=_iso_week_label(current),
                start_date=current,
                end_date=period_end,
            )
        )
        current += timedelta(days=7)
    return periods


def _section_value(section) -> str:
    return getattr(section, "value", section) or "UNCLASSIFIED"


async def _build_cashflow_statement(
    db: AsyncSession,
    *,
    granularity: str,
    date_from: date,
    date_to: date,
    company_id: uuid.UUID | None,
    include_intercompany: bool,
) -> CashFlowStatement:
    periods = _periods_between(date_from, date_to, granularity)
    period_index = {p.key: idx for idx, p in enumerate(periods)}
    period_count = len(periods)

    period_expr = func.date_trunc("month" if granularity == "monthly" else "week", Movement.value_date).label("period_start")

    stmt = (
        select(
            period_expr,
            CategoryTaxonomy.cash_flow_section.label("section"),
            func.coalesce(MovementClassification.category_code, "UNCLASSIFIED").label("category_code"),
            func.coalesce(CategoryTaxonomy.name, "Unclassified").label("category_name"),
            func.sum(Movement.amount).label("amount"),
        )
        .select_from(Movement)
        .outerjoin(MovementClassification, Movement.id == MovementClassification.movement_id)
        .outerjoin(CategoryTaxonomy, MovementClassification.category_code == CategoryTaxonomy.code)
        .where(
            Movement.is_deleted.is_(False),
            Movement.value_date >= date_from,
            Movement.value_date <= date_to,
        )
        .group_by(text("period_start"), CategoryTaxonomy.cash_flow_section, MovementClassification.category_code, CategoryTaxonomy.name)
    )
    if company_id:
        stmt = stmt.where(Movement.company_id == company_id)
    if not include_intercompany:
        stmt = stmt.where(Movement.is_intercompany.is_(False))

    rows = (await db.execute(stmt)).all()

    row_map: dict[tuple[str, str], dict] = {}
    section_order = {
        "OPERATING": 0,
        "INVESTING": 1,
        "FINANCING": 2,
        "INTERNAL": 3,
        "UNCLASSIFIED": 4,
    }
    section_totals: dict[str, list[Decimal]] = {
        section: [Decimal(0) for _ in range(period_count)]
        for section in section_order
    }

    for row in rows:
        period_start = row.period_start.date() if hasattr(row.period_start, "date") else row.period_start
        key = period_start.isoformat()
        if key not in period_index:
            continue
        idx = period_index[key]
        section = _section_value(row.section)
        category_code = row.category_code or "UNCLASSIFIED"
        category_name = row.category_name or "Unclassified"
        amount = row.amount or Decimal(0)

        map_key = (section, category_code)
        if map_key not in row_map:
            row_map[map_key] = {
                "section": section,
                "category_code": category_code,
                "category_name": category_name,
                "values": [Decimal(0) for _ in range(period_count)],
            }
        row_map[map_key]["values"][idx] += amount
        section_totals.setdefault(section, [Decimal(0) for _ in range(period_count)])
        section_totals[section][idx] += amount

    cashflow_rows = [
        CashFlowRow(
            section=data["section"],
            category_code=data["category_code"],
            category_name=data["category_name"],
            values=data["values"],
            total=sum(data["values"], Decimal(0)),
        )
        for data in row_map.values()
    ]
    cashflow_rows.sort(key=lambda r: (section_order.get(r.section, 99), r.category_name, r.category_code))

    section_summaries = [
        CashFlowSectionSummary(
            section=section,
            values=values,
            total=sum(values, Decimal(0)),
        )
        for section, values in sorted(section_totals.items(), key=lambda item: section_order.get(item[0], 99))
        if any(v != 0 for v in values)
    ]

    net_cash_flow = [Decimal(0) for _ in range(period_count)]
    for values in section_totals.values():
        for idx, value in enumerate(values):
            net_cash_flow[idx] += value

    return CashFlowStatement(
        granularity=granularity,
        date_from=date_from,
        date_to=date_to,
        company_id=company_id,
        include_intercompany=include_intercompany,
        periods=periods,
        sections=section_summaries,
        rows=cashflow_rows,
        net_cash_flow=net_cash_flow,
        net_cash_flow_total=sum(net_cash_flow, Decimal(0)),
        as_of=date.today(),
    )


@router.get("/dashboard", response_model=DashboardSummary)
async def get_dashboard(db: AsyncSession = Depends(get_db)):
    today = date.today()

    # ── 1. Cash position per company ──────────────────────────────────────────
    # Step A: latest balance_after per bank account (most recent movement with non-NULL balance)
    latest_balance_subq = (
        select(
            Movement.bank_account_id,
            func.max(Movement.value_date).label("max_date"),
        )
        .where(
            Movement.is_deleted.is_(False),
            Movement.balance_after.is_not(None),
        )
        .group_by(Movement.bank_account_id)
        .subquery()
    )

    balance_rows = (
        await db.execute(
            select(
                Movement.bank_account_id,
                Movement.balance_after,
            )
            .join(
                latest_balance_subq,
                (Movement.bank_account_id == latest_balance_subq.c.bank_account_id)
                & (Movement.value_date == latest_balance_subq.c.max_date),
            )
            .where(Movement.is_deleted.is_(False), Movement.balance_after.is_not(None))
        )
    ).all()

    # Map bank_account_id → last known balance
    account_balance: dict[str, Decimal] = {}
    for row in balance_rows:
        aid = str(row.bank_account_id)
        # In case multiple movements on the same max_date, keep the first
        if aid not in account_balance:
            account_balance[aid] = row.balance_after

    # Step B: net flow per bank account (SUM of all amounts)
    net_flow_rows = (
        await db.execute(
            select(
                Movement.bank_account_id,
                func.sum(Movement.amount).label("net"),
            )
            .where(Movement.is_deleted.is_(False))
            .group_by(Movement.bank_account_id)
        )
    ).all()
    account_net: dict[str, Decimal] = {str(r.bank_account_id): r.net or Decimal(0) for r in net_flow_rows}

    # Step C: aggregate by company
    company_accounts = (
        await db.execute(
            select(
                BankAccount.id,
                BankAccount.company_id,
                Company.name.label("company_name"),
                Company.short_name,
            )
            .join(Company, BankAccount.company_id == Company.id)
            .where(BankAccount.is_active.is_(True))
        )
    ).all()

    company_data: dict[str, dict] = {}
    for row in company_accounts:
        cid = str(row.company_id)
        aid = str(row.id)
        if cid not in company_data:
            company_data[cid] = {
                "company_id": row.company_id,
                "company_name": row.company_name,
                "short_name": row.short_name,
                "last_balance": Decimal(0),
                "net_flow": Decimal(0),
                "has_balance_data": False,
            }
        if aid in account_balance:
            company_data[cid]["last_balance"] += account_balance[aid]
            company_data[cid]["has_balance_data"] = True
        company_data[cid]["net_flow"] += account_net.get(aid, Decimal(0))

    cash_by_company: list[CompanyCashPosition] = []
    for cid, d in company_data.items():
        cash_by_company.append(
            CompanyCashPosition(
                company_id=d["company_id"],
                company_name=d["company_name"],
                short_name=d["short_name"],
                last_balance=d["last_balance"] if d["has_balance_data"] else None,
                net_flow=d["net_flow"],
                has_balance_data=d["has_balance_data"],
            )
        )

    # Total cash: prefer last_balance when available, else net_flow
    total_cash = sum(
        (c.last_balance if c.last_balance is not None else c.net_flow) for c in cash_by_company
    ) or Decimal(0)

    # ── 2. Net flow week-to-date ──────────────────────────────────────────────
    week_start = _last_monday(today)
    wtd_result = await db.execute(
        select(func.sum(Movement.amount))
        .where(
            Movement.is_deleted.is_(False),
            Movement.value_date >= week_start,
            Movement.value_date <= today,
        )
    )
    net_flow_wtd: Decimal = wtd_result.scalar_one() or Decimal(0)

    # ── 3. Intercompany match counts ──────────────────────────────────────────
    ic_counts = (
        await db.execute(
            select(
                IntercompanyMatch.status,
                func.count(IntercompanyMatch.id).label("cnt"),
            ).group_by(IntercompanyMatch.status)
        )
    ).all()

    status_map = {str(r.status): r.cnt for r in ic_counts}
    pending_ic = status_map.get(MatchStatus.PROPOSED.value, 0)
    in_transit_ic = status_map.get(MatchStatus.IN_TRANSIT.value, 0)
    unresolved_ic = status_map.get(MatchStatus.UNRESOLVED.value, 0)

    # ── 4. Weekly cash flow — last 13 ISO weeks ───────────────────────────────
    thirteen_weeks_ago = today - timedelta(weeks=13)

    weekly_rows = (
        await db.execute(
            select(
                func.date_trunc("week", Movement.value_date).label("week_start"),
                func.sum(
                    case((Movement.amount > 0, Movement.amount), else_=Decimal(0))
                ).label("inflow"),
                func.sum(
                    case((Movement.amount < 0, Movement.amount), else_=Decimal(0))
                ).label("outflow"),
                func.sum(Movement.amount).label("net"),
            )
            .where(
                Movement.is_deleted.is_(False),
                Movement.value_date >= thirteen_weeks_ago,
                Movement.value_date <= today,
            )
            .group_by(text("week_start"))
            .order_by(text("week_start"))
        )
    ).all()

    weekly_cash_flow: list[WeeklyCashFlow] = [
        WeeklyCashFlow(
            week_start=row.week_start.date() if hasattr(row.week_start, "date") else row.week_start,
            week_label=_iso_week_label(
                row.week_start.date() if hasattr(row.week_start, "date") else row.week_start
            ),
            inflow=row.inflow or Decimal(0),
            outflow=row.outflow or Decimal(0),
            net=row.net or Decimal(0),
        )
        for row in weekly_rows
    ]

    return DashboardSummary(
        cash_by_company=cash_by_company,
        total_cash=total_cash,
        net_flow_wtd=net_flow_wtd,
        pending_ic_matches=pending_ic,
        in_transit_ic=in_transit_ic,
        unresolved_ic=unresolved_ic,
        weekly_cash_flow=weekly_cash_flow,
        as_of=today,
    )


@router.get("/cashflow", response_model=CashFlowStatement)
async def get_cashflow_statement(
    granularity: str = Query("weekly", pattern="^(weekly|monthly)$"),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    company_id: uuid.UUID | None = Query(None),
    include_intercompany: bool = Query(False),
    db: AsyncSession = Depends(get_db),
):
    today = date.today()
    if date_to is None:
        date_to = today
    if date_from is None:
        date_from = _month_start(today) if granularity == "monthly" else today - timedelta(weeks=13)
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    return await _build_cashflow_statement(
        db,
        granularity=granularity,
        date_from=date_from,
        date_to=date_to,
        company_id=company_id,
        include_intercompany=include_intercompany,
    )


@router.get("/cashflow/export")
async def export_cashflow_statement(
    granularity: str = Query("weekly", pattern="^(weekly|monthly)$"),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    company_id: uuid.UUID | None = Query(None),
    include_intercompany: bool = Query(False),
    db: AsyncSession = Depends(get_db),
):
    today = date.today()
    if date_to is None:
        date_to = today
    if date_from is None:
        date_from = _month_start(today) if granularity == "monthly" else today - timedelta(weeks=13)
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    statement = await _build_cashflow_statement(
        db,
        granularity=granularity,
        date_from=date_from,
        date_to=date_to,
        company_id=company_id,
        include_intercompany=include_intercompany,
    )

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Statement"

    ws.append(["Cash Flow Statement"])
    ws.append(["Granularity", statement.granularity])
    ws.append(["Period", f"{statement.date_from.isoformat()} to {statement.date_to.isoformat()}"])
    ws.append(["Intercompany included", "Yes" if statement.include_intercompany else "No"])
    ws.append([])

    headers = ["Section", "Category Code", "Category"] + [p.label for p in statement.periods] + ["Total"]
    ws.append(headers)
    header_row = ws.max_row

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    current_section = None
    for row in statement.rows:
        if row.section != current_section:
            summary = next((s for s in statement.sections if s.section == row.section), None)
            if summary:
                ws.append([row.section, "", f"Total {row.section}"] + [float(v) for v in summary.values] + [float(summary.total)])
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
            current_section = row.section
        ws.append([row.section, row.category_code, row.category_name] + [float(v) for v in row.values] + [float(row.total)])

    ws.append([])
    ws.append(["", "", "NET CASH FLOW"] + [float(v) for v in statement.net_cash_flow] + [float(statement.net_cash_flow_total)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row[3:]:
            cell.number_format = '#,##0.00;[Red](#,##0.00)'

    ws.freeze_panes = f"D{header_row + 1}"
    for col_idx in range(1, ws.max_column + 1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 36)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"cashflow_{granularity}_{statement.date_from.isoformat()}_{statement.date_to.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/consistency", response_model=ConsistencyReport)
async def get_consistency(
    year: int = Query(default=None),
    month: int = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    today = date.today()
    if year is None:
        year = today.year
    if month is None:
        month = today.month

    period_start = date(year, month, 1)
    period_end = date(year + 1, 1, 1) - timedelta(days=1) if month == 12 else date(year, month + 1, 1) - timedelta(days=1)
    period_label = period_start.strftime("%B %Y")

    # ── Section A: Import Completeness ────────────────────────────────────────
    accounts = (
        await db.execute(
            select(
                BankAccount.id,
                BankAccount.account_name,
                BankAccount.bank_name,
                BankAccount.iban,
                BankAccount.company_id,
                Company.name.label("company_name"),
                Company.short_name,
            )
            .join(Company, BankAccount.company_id == Company.id)
            .where(BankAccount.is_active.is_(True))
        )
    ).all()

    movement_stats = (
        await db.execute(
            select(
                Movement.bank_account_id,
                func.count(Movement.id).label("cnt"),
                func.min(Movement.value_date).label("earliest"),
                func.max(Movement.value_date).label("latest"),
            )
            .where(
                Movement.is_deleted.is_(False),
                Movement.value_date >= period_start,
                Movement.value_date <= period_end,
            )
            .group_by(Movement.bank_account_id)
        )
    ).all()

    movement_stat_map = {str(r.bank_account_id): r for r in movement_stats}

    last_batch_rows = (
        await db.execute(
            select(
                ImportBatch.bank_account_id,
                func.max(ImportBatch.imported_at).label("last_at"),
            )
            .where(ImportBatch.status == ImportStatus.COMPLETED)
            .group_by(ImportBatch.bank_account_id)
        )
    ).all()

    last_batch_map = {str(r.bank_account_id): r.last_at for r in last_batch_rows}

    section_a: list[AccountImportStatus] = []
    for acc in accounts:
        aid = str(acc.id)
        stats = movement_stat_map.get(aid)
        if stats is None or stats.cnt == 0:
            coverage_status = ImportCoverageStatus.MISSING
            earliest = None
            latest = None
            cnt = 0
        else:
            cnt = stats.cnt
            earliest = stats.earliest
            latest = stats.latest
            coverage_status = (
                ImportCoverageStatus.PARTIAL
                if (period_end - latest).days > 3
                else ImportCoverageStatus.OK
            )
        section_a.append(
            AccountImportStatus(
                bank_account_id=acc.id,
                account_name=acc.account_name,
                bank_name=acc.bank_name,
                company_name=acc.company_name,
                short_name=acc.short_name,
                iban_last4=acc.iban[-4:] if acc.iban else "----",
                movement_count=cnt,
                earliest_movement=earliest,
                latest_movement=latest,
                last_batch_at=last_batch_map.get(aid),
                status=coverage_status,
            )
        )

    _status_order = {ImportCoverageStatus.MISSING: 0, ImportCoverageStatus.PARTIAL: 1, ImportCoverageStatus.OK: 2}
    section_a.sort(key=lambda x: (_status_order[x.status], x.company_name, x.account_name))

    # ── Section B: Balance Reconciliation ─────────────────────────────────────
    accounts_with_balance_ids = (
        await db.execute(
            select(Movement.bank_account_id)
            .where(Movement.is_deleted.is_(False), Movement.balance_after.is_not(None))
            .group_by(Movement.bank_account_id)
        )
    ).scalars().all()

    acc_map = {str(a.id): a for a in accounts}
    section_b: list[BalanceReconciliation] = []

    for aid_uuid in accounts_with_balance_ids:
        aid_str = str(aid_uuid)
        acc = acc_map.get(aid_str)
        if not acc:
            continue

        opening_balance: Decimal | None = (
            await db.execute(
                select(Movement.balance_after)
                .where(
                    Movement.bank_account_id == aid_uuid,
                    Movement.is_deleted.is_(False),
                    Movement.balance_after.is_not(None),
                    Movement.value_date < period_start,
                )
                .order_by(Movement.value_date.desc())
                .limit(1)
            )
        ).scalar_one_or_none()

        closing_balance_bank: Decimal | None = (
            await db.execute(
                select(Movement.balance_after)
                .where(
                    Movement.bank_account_id == aid_uuid,
                    Movement.is_deleted.is_(False),
                    Movement.balance_after.is_not(None),
                    Movement.value_date >= period_start,
                    Movement.value_date <= period_end,
                )
                .order_by(Movement.value_date.desc())
                .limit(1)
            )
        ).scalar_one_or_none()

        net_in_period: Decimal = (
            await db.execute(
                select(func.sum(Movement.amount))
                .where(
                    Movement.bank_account_id == aid_uuid,
                    Movement.is_deleted.is_(False),
                    Movement.value_date >= period_start,
                    Movement.value_date <= period_end,
                )
            )
        ).scalar_one() or Decimal(0)

        if closing_balance_bank is None:
            recon_status = "NO_DATA"
            delta = None
            closing_balance_computed = None
        else:
            opening = opening_balance if opening_balance is not None else Decimal(0)
            closing_balance_computed = opening + net_in_period
            delta = closing_balance_bank - closing_balance_computed
            abs_delta = abs(delta)
            if abs_delta <= Decimal("0.01"):
                recon_status = "OK"
            elif abs_delta <= Decimal("50"):
                recon_status = "WARNING"
            else:
                recon_status = "ERROR"

        section_b.append(
            BalanceReconciliation(
                bank_account_id=aid_uuid,
                account_name=acc.account_name,
                bank_name=acc.bank_name,
                company_name=acc.company_name,
                period_label=period_label,
                opening_balance=opening_balance,
                closing_balance_bank=closing_balance_bank,
                closing_balance_computed=closing_balance_computed,
                delta=delta,
                status=recon_status,
            )
        )

    _recon_order = {"ERROR": 0, "WARNING": 1, "NO_DATA": 2, "OK": 3}
    section_b.sort(key=lambda x: (_recon_order.get(x.status, 99), x.company_name, x.account_name))

    # ── Section C1: HoldCo revenue warnings ───────────────────────────────────
    holdco_rows = (
        await db.execute(
            select(
                Movement.id,
                Movement.value_date,
                Movement.amount,
                Movement.description,
                BankAccount.account_name,
                Company.name.label("company_name"),
            )
            .join(BankAccount, Movement.bank_account_id == BankAccount.id)
            .join(Company, BankAccount.company_id == Company.id)
            .join(MovementClassification, MovementClassification.movement_id == Movement.id)
            .where(
                Company.is_holding.is_(True),
                MovementClassification.category_code == "OCF_INCOME",
                Movement.is_deleted.is_(False),
            )
        )
    ).all()

    holdco_revenue_warnings = [
        DataQualityWarning(
            rule="HOLDCO_REVENUE",
            company_name=r.company_name,
            account_name=r.account_name,
            movement_id=r.id,
            movement_date=r.value_date,
            movement_amount=r.amount,
            description=r.description,
        )
        for r in holdco_rows
    ]

    # ── Section C2: High unclassified rate ────────────────────────────────────
    total_per_company = (
        await db.execute(
            select(
                Movement.company_id,
                Company.name.label("company_name"),
                func.count(Movement.id).label("total"),
            )
            .join(Company, Movement.company_id == Company.id)
            .where(
                Movement.is_deleted.is_(False),
                Movement.value_date >= period_start,
                Movement.value_date <= period_end,
            )
            .group_by(Movement.company_id, Company.name)
        )
    ).all()

    unclassified_per_company = (
        await db.execute(
            select(
                Movement.company_id,
                func.count(Movement.id).label("unclassified"),
            )
            .join(MovementClassification, MovementClassification.movement_id == Movement.id)
            .where(
                Movement.is_deleted.is_(False),
                Movement.value_date >= period_start,
                Movement.value_date <= period_end,
                MovementClassification.category_code == "UNCLASSIFIED",
            )
            .group_by(Movement.company_id)
        )
    ).all()

    unclassified_map = {str(r.company_id): r.unclassified for r in unclassified_per_company}

    high_unclassified_companies = [
        UnclassifiedRateWarning(
            company_name=row.company_name,
            total_movements=row.total,
            unclassified_count=unclassified_map.get(str(row.company_id), 0),
            unclassified_rate=unclassified_map.get(str(row.company_id), 0) / row.total,
        )
        for row in total_per_company
        if row.total > 0 and unclassified_map.get(str(row.company_id), 0) / row.total > 0.15
    ]

    # ── Section C3 & C4: IC counts ────────────────────────────────────────────
    now_utc = datetime.now(timezone.utc)

    unresolved_ic_count: int = (
        await db.execute(
            select(func.count(IntercompanyMatch.id))
            .where(IntercompanyMatch.status == MatchStatus.UNRESOLVED)
        )
    ).scalar_one() or 0

    in_transit_timeout_count: int = (
        await db.execute(
            select(func.count(IntercompanyMatch.id))
            .where(
                IntercompanyMatch.status == MatchStatus.IN_TRANSIT,
                IntercompanyMatch.transit_expires_at < now_utc,
            )
        )
    ).scalar_one() or 0

    return ConsistencyReport(
        period_year=year,
        period_month=month,
        period_label=period_label,
        section_a=section_a,
        section_b=section_b,
        holdco_revenue_warnings=holdco_revenue_warnings,
        high_unclassified_companies=high_unclassified_companies,
        unresolved_ic_count=unresolved_ic_count,
        in_transit_timeout_count=in_transit_timeout_count,
        as_of=today,
    )
